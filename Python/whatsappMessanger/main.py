import pywhatkit
import openpyxl
import pandas as pd
import time

wb = openpyxl.load_workbook("data.xlsx")
sh = wb.active

for i in range(2, sh.max_row + 1):
    c1 = sh.cell(row=i, column=4)
    # print(c1.value, c2.value)
    pywhatkit.sendwhatmsg_instantly(
        '+91' + str(c1.value), "Omkar Date", 5, True)
    time.sleep(30)

# pywhatkit.sendwhatmsg('+91' + str(c1.value),
#                           "Omkar Date", 1, 27, 3, True, 1)
